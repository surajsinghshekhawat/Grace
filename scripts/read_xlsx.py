import sys
import os

try:
    import openpyxl
    # Run from project root: python scripts/read_xlsx.py
    _root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    _xlsx = os.path.join(_root, "data", "raw", "RV PhD Data Sheet 15.03.2022 (1).xlsx")
    wb = openpyxl.load_workbook(_xlsx, read_only=True, data_only=True)
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        nrows = len(rows)
        ncols = max(len(r) for r in rows) if rows else 0
        print("Sheet:", repr(name))
        print("  Total rows:", nrows, "| Data readings (excl. header rows):", nrows - 3)
        print("  Columns:", ncols)
        # Row 3 = header with variable names/numbers
        h = list(rows[2]) if len(rows) > 2 else []
        print("  Header row (all):", [str(x)[:25] for x in h])
        # First column of first 5 data rows (to see ID pattern)
        if nrows > 4:
            col0 = [rows[i][0] for i in range(3, min(8, nrows))]
            print("  First col (sample):", col0)
        print("---")
    wb.close()
except Exception as e:
    print("Error:", e)
    import traceback
    traceback.print_exc()
    sys.exit(1)
